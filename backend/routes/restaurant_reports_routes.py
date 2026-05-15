from flask import Blueprint, request, jsonify
from psycopg2.extras import RealDictCursor
from db import get_db_connection
import jwt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from flask import send_file
import io
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Table, TableStyle, Spacer
)
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.units import inch

JWT_SECRET = "MAHAL_SUPER_SECRET_2025"

restaurant_reports_bp = Blueprint("restaurant_reports_bp", __name__)


# =========================================================
# AUTH
# =========================================================
def get_restaurant_from_token():

    auth = request.headers.get("Authorization", "")

    if not auth.startswith("Bearer "):
        return None, ("Unauthorized", 401)

    try:

        decoded = jwt.decode(
            auth.replace("Bearer ", ""),
            JWT_SECRET,
            algorithms=["HS256"]
        )

        if decoded.get("role", "").upper() != "RESTAURANT":
            return None, ("Forbidden", 403)

        return decoded.get("linked_id"), None

    except Exception:
        return None, ("Invalid token", 401)

@restaurant_reports_bp.route(
    "/restaurant/reports/purchases",
    methods=["GET"]
)
def purchase_report():

    restaurant_id, err = get_restaurant_from_token()
    if err:
        return jsonify([]), 200

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")

        query = """
            SELECT
                oh.order_id,
                oh.supplier_id,
                oh.order_date,
                oh.expected_delivery_date,
                oh.total_amount,
                oh.status,
                oh.payment_status,
                oh.payment_method,
                oh.restaurant_paid_amount,
                oh.restaurant_due_amount,
                oh.remarks,

                sr.company_name_english,
                sr.company_name_arabic,

                oi.product_id,
                oi.product_name_english,
                pm.product_name_arabic,
                oi.quantity,
                oi.price_per_unit,
                oi.discount,
                oi.total_amount AS item_total

            FROM order_header oh

            JOIN order_items oi
                ON oi.order_id = oh.order_id

            JOIN supplier_registration sr
                ON sr.supplier_id = oh.supplier_id

            LEFT JOIN product_management pm
                ON pm.product_id = oi.product_id

            WHERE oh.restaurant_id = %s
        """

        params = [restaurant_id]

        if start_date and end_date:
            query += " AND DATE(oh.order_date) BETWEEN %s AND %s"
            params.extend([start_date, end_date])

        query += " ORDER BY oh.order_date DESC"

        cur.execute(query, tuple(params))
        rows = cur.fetchall()

        # 🔥 GROUP BY ORDER
        orders_map = {}

        for r in rows:
            oid = r["order_id"]

            if oid not in orders_map:
                orders_map[oid] = {
                    "order_id": oid,
                    "supplier_id": r["supplier_id"],
                    "order_date": r["order_date"],
                    "expected_delivery_date": r["expected_delivery_date"],
                    "total_amount": r["total_amount"],
                    "status": r["status"],
                    "payment_status": r["payment_status"],
                    "payment_method": r["payment_method"],
                    "restaurant_paid_amount": r["restaurant_paid_amount"],
                    "restaurant_due_amount": r["restaurant_due_amount"],
                    "remarks": r["remarks"],
                    "company_name_english": r["company_name_english"],
                    "company_name_arabic": r["company_name_arabic"],
                    "items": []
                }

            # ADD ITEMS
            orders_map[oid]["items"].append({
                "product_id": r["product_id"],
                "product_name_english": r["product_name_english"],
                "product_name_arabic": r["product_name_arabic"],
                "quantity": r["quantity"],
                "price_per_unit": r["price_per_unit"],
                "discount": r["discount"],
                "total_amount": r["item_total"]
            })

        result = list(orders_map.values())

        return jsonify(result), 200

    finally:
        cur.close()
        conn.close()
# =========================================================
# GRN REPORT  ✅ FIXED (ONE ROW PER GRN)
# =========================================================
@restaurant_reports_bp.route(
    "/restaurant/reports/grn",
    methods=["GET"]
)
def grn_report():

    restaurant_id, err = get_restaurant_from_token()
    if err:
        return jsonify([]), 200

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:

        cur.execute("""
            SELECT
                gh.grn_id,
                gh.order_id,
                gh.status,
                gh.created_at,

                sr.company_name_english,
                sr.company_name_arabic,

                COALESCE(SUM(gi.received_quantity), 0) AS received_qty,
                COUNT(gi.grn_item_id) AS total_items

            FROM grn_header gh

            JOIN supplier_registration sr
                ON sr.supplier_id = gh.supplier_id

            LEFT JOIN grn_items gi
                ON gi.grn_id = gh.grn_id

            WHERE gh.restaurant_id = %s

            GROUP BY
                gh.grn_id,
                gh.order_id,
                gh.status,
                gh.created_at,
                sr.company_name_english,
                sr.company_name_arabic

            ORDER BY gh.created_at DESC
        """, (restaurant_id,))

        rows = cur.fetchall()

        return jsonify(rows), 200

    finally:
        cur.close()
        conn.close()


# =========================================================
# SUPPLIER PERFORMANCE REPORT
# =========================================================
@restaurant_reports_bp.route(
    "/restaurant/reports/suppliers",
    methods=["GET"]
)
def supplier_report():

    restaurant_id, err = get_restaurant_from_token()
    if err:
        return jsonify([]), 200

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:

        cur.execute("""
            SELECT
                sr.supplier_id,
                sr.company_name_english,
                sr.company_name_arabic,

                COUNT(DISTINCT oh.order_id) AS total_orders,
                COALESCE(SUM(oh.total_amount), 0) AS total_purchase,

                COUNT(
                    CASE WHEN oh.status = 'DELIVERED' THEN 1 END
                ) AS delivered_orders,

                COUNT(
                    CASE WHEN oh.status != 'DELIVERED' THEN 1 END
                ) AS pending_orders

            FROM supplier_registration sr

            JOIN order_header oh
                ON oh.supplier_id = sr.supplier_id

            WHERE oh.restaurant_id = %s

            GROUP BY sr.supplier_id, sr.company_name_english

            ORDER BY total_purchase DESC
        """, (restaurant_id,))

        return jsonify(cur.fetchall()), 200

    finally:
        cur.close()
        conn.close()
# =========================================================
# ENTERPRISE INVOICE REPORT
# =========================================================
@restaurant_reports_bp.route(
    "/restaurant/reports/invoices",
    methods=["GET"]
)
def invoice_report():

    restaurant_id, err = get_restaurant_from_token()

    if err:
        return jsonify([]), 200

    conn = get_db_connection()

    cur = conn.cursor(
        cursor_factory=RealDictCursor
    )

    try:

        cur.execute("""

            SELECT
                ih.invoice_id,
                ih.invoice_number,
                ih.order_id,

                ih.invoice_date,
                ih.invoice_status,
                ih.payment_status,

                ih.subtotal_amount,
                ih.discount_amount,
                ih.tax_amount,
                ih.grand_total,

                ih.created_at,

                sr.company_name_english
                    AS supplier_name,

                sr.company_name_arabic
                    AS supplier_name_arabic,

                COUNT(
                    ii.invoice_item_id
                ) AS total_items,

                COALESCE(
                    SUM(ii.quantity),
                    0
                ) AS total_quantity

            FROM invoice_header ih

            JOIN supplier_registration sr
                ON sr.supplier_id =
                    ih.supplier_id

            LEFT JOIN invoice_items ii
                ON ii.invoice_id =
                    ih.invoice_id

            WHERE ih.restaurant_id = %s

            GROUP BY
                ih.invoice_id,
                ih.invoice_number,
                ih.order_id,
                ih.invoice_date,
                ih.invoice_status,
                ih.payment_status,
                ih.subtotal_amount,
                ih.discount_amount,
                ih.tax_amount,
                ih.grand_total,
                ih.created_at,
                sr.company_name_english,
                sr.company_name_arabic

            ORDER BY
                ih.invoice_date DESC

        """, (restaurant_id,))

        rows = cur.fetchall()

        return jsonify(rows), 200

    finally:

        cur.close()
        conn.close()
@restaurant_reports_bp.route("/restaurant/reports/purchases/export", methods=["GET"])
def export_purchase_report():

    restaurant_id, err = get_restaurant_from_token()
    if err:
        return jsonify({"error": err[0]}), err[1]

    status = request.args.get("status")
    supplier = request.args.get("supplier")

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    # ✅ STEP 1: BASE QUERY
    query = """
        SELECT 
            oh.order_id,
            oh.order_date,
            oh.status,
            sr.company_name_english AS supplier,
            sr.company_name_arabic,
            oi.product_name_english,
            oi.quantity,
            oi.total_amount
        FROM order_header oh
        JOIN order_items oi ON oi.order_id = oh.order_id
        JOIN supplier_registration sr ON sr.supplier_id = oh.supplier_id
        WHERE oh.restaurant_id = %s
    """

    params = [restaurant_id]

    # ✅ STEP 2: APPLY FILTERS
    if status and status != "ALL":
        query += " AND oh.status = %s"
        params.append(status)

    if supplier and supplier != "ALL":
        query += """
            AND (
                sr.company_name_english = %s
                OR sr.company_name_arabic = %s
            )
        """
        params.extend([supplier, supplier])

    query += " ORDER BY oh.order_date DESC"

    # ✅ STEP 3: EXECUTE
    cur.execute(query, tuple(params))

    # ✅ STEP 4: FETCH
    rows = cur.fetchall()

    # ✅ BUILD EXCEL
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Report"

    # headers
    ws.append(["Order ID", "Date", "Supplier", "Product", "Qty", "Amount"])

    for r in rows:
        ws.append([
            r["order_id"],
            str(r["order_date"]),
            r["supplier"],
            r["product_name_english"],
            r["quantity"],
            r["total_amount"]
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        download_name="purchase_report.xlsx",
        as_attachment=True
    )

    

@restaurant_reports_bp.route("/restaurant/reports/purchases/pdf", methods=["GET"])
def export_purchase_pdf():

    restaurant_id, err = get_restaurant_from_token()
    if err:
        return jsonify({"error": err[0]}), err[1]

    status = request.args.get("status")
    supplier = request.args.get("supplier")

    conn = get_db_connection()
    cur = conn.cursor(cursor_factory=RealDictCursor)

    # ✅ DEFINE QUERY FIRST
    query = """
        SELECT 
            oh.order_id,
            oh.order_date,
            oh.status,
            sr.company_name_english,
            sr.company_name_arabic,
            oi.product_name_english,
            oi.quantity,
            oi.total_amount
        FROM order_header oh
        JOIN order_items oi ON oi.order_id = oh.order_id
        JOIN supplier_registration sr ON sr.supplier_id = oh.supplier_id
        WHERE oh.restaurant_id = %s
    """

    params = [restaurant_id]

    # ✅ APPLY FILTERS BEFORE EXECUTE
    if status and status != "ALL":
        query += " AND oh.status = %s"
        params.append(status)

    if supplier and supplier != "ALL":
        query += " AND sr.company_name_english = %s"
        params.append(supplier)

    query += " ORDER BY oh.order_date DESC"

    # ✅ EXECUTE ONLY ONCE
    cur.execute(query, tuple(params))
    rows = cur.fetchall()

    # ---------------- PDF BUILD ----------------
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)

    styles = getSampleStyleSheet()
    elements = []

    # HEADER
    elements.append(Paragraph("<b>MAHAL</b>", styles["Title"]))
    elements.append(Spacer(1, 6))
    elements.append(Paragraph("Purchase Report", styles["Heading2"]))
    elements.append(Spacer(1, 12))

    # TABLE
    data = [["Order ID", "Date", "Supplier", "Product", "Qty", "Amount"]]

    for r in rows:
        data.append([
            r["order_id"],
            str(r["order_date"])[:10],
            r["company_name_english"],
            r["product_name_english"],
            r["quantity"],
            f"{r['total_amount']:.2f}"
        ])

    table = Table(
        data,
        colWidths=[1.4*inch, 1.1*inch, 2*inch, 2*inch, 0.6*inch, 1*inch]
    )

    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.orange),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (4, 1), (5, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
    ]))

    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        download_name="purchase_report.pdf",
        as_attachment=True
    )

# =========================================================
# GRN EXCEL EXPORT
# =========================================================
@restaurant_reports_bp.route(
    "/restaurant/reports/grn/export",
    methods=["GET"]
)
def export_grn_excel():

    restaurant_id, err = get_restaurant_from_token()

    if err:
        return jsonify({"error": err[0]}), err[1]

    status = request.args.get("status")
    supplier = request.args.get("supplier")
    search = request.args.get("search")

    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    conn = get_db_connection()

    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:

        # =====================================================
        # QUERY
        # =====================================================
        query = """
            SELECT
                gh.grn_id,
                gh.order_id,
                gh.status,
                gh.created_at,

                sr.company_name_english,
                sr.company_name_arabic,

                COALESCE(
                    SUM(gi.received_quantity),
                    0
                ) AS received_qty,

                COUNT(gi.grn_item_id) AS total_items

            FROM grn_header gh

            JOIN supplier_registration sr
                ON sr.supplier_id = gh.supplier_id

            LEFT JOIN grn_items gi
                ON gi.grn_id = gh.grn_id

            WHERE gh.restaurant_id = %s
        """

        params = [restaurant_id]

        # =====================================================
        # STATUS FILTER
        # =====================================================
        if status and status != "ALL":

            query += """
                AND gh.status = %s
            """

            params.append(status)

        # =====================================================
        # SUPPLIER FILTER
        # =====================================================
        if supplier and supplier != "ALL":

            query += """
                AND (
                    sr.company_name_english = %s
                    OR sr.company_name_arabic = %s
                )
            """

            params.extend([
                supplier,
                supplier
            ])

        # =====================================================
        # SEARCH FILTER
        # =====================================================
        if search:

            query += """
                AND (
                    CAST(gh.grn_id AS TEXT)
                    ILIKE %s

                    OR CAST(gh.order_id AS TEXT)
                    ILIKE %s

                    OR sr.company_name_english
                    ILIKE %s

                    OR sr.company_name_arabic
                    ILIKE %s
                )
            """

            params.extend([
                f"%{search}%",
                f"%{search}%",
                f"%{search}%",
                f"%{search}%"
            ])

        # =====================================================
        # DATE FILTER
        # =====================================================
        if start_date and end_date:

            query += """
                AND DATE(gh.created_at)
                BETWEEN %s AND %s
            """

            params.extend([
                start_date,
                end_date
            ])

        # =====================================================
        # GROUP + SORT
        # =====================================================
        query += """

            GROUP BY
                gh.grn_id,
                gh.order_id,
                gh.status,
                gh.created_at,
                sr.company_name_english,
                sr.company_name_arabic

            ORDER BY gh.created_at DESC
        """

        # =====================================================
        # EXECUTE
        # =====================================================
        cur.execute(query, tuple(params))

        rows = cur.fetchall()

        # =====================================================
        # CREATE EXCEL
        # =====================================================
        wb = Workbook()

        ws = wb.active

        ws.title = "GRN Report"

        # =====================================================
        # HEADERS
        # =====================================================
        headers = [
            "GRN ID",
            "ORDER ID",
            "SUPPLIER",
            "STATUS",
            "RECEIVED QTY",
            "TOTAL ITEMS",
            "DATE"
        ]

        ws.append(headers)

        # =====================================================
        # HEADER STYLING
        # =====================================================
        for cell in ws[1]:

            cell.font = Font(
                bold=True,
                size=12
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center"
            )

        # =====================================================
        # DATA ROWS
        # =====================================================
        for r in rows:

            ws.append([
                r["grn_id"],
                r["order_id"],
                r["company_name_english"],
                r["status"],
                r["received_qty"],
                r["total_items"],
                str(r["created_at"])[:10]
            ])

        # =====================================================
        # COLUMN WIDTHS
        # =====================================================
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 18

        # =====================================================
        # SAVE FILE
        # =====================================================
        output = io.BytesIO()

        wb.save(output)

        output.seek(0)

        return send_file(
            output,
            download_name="grn_report.xlsx",
            as_attachment=True
        )

    finally:

        cur.close()
        conn.close()


# =========================================================
# GRN PDF EXPORT
# =========================================================
@restaurant_reports_bp.route(
    "/restaurant/reports/grn/pdf",
    methods=["GET"]
)
def export_grn_pdf():

    restaurant_id, err = get_restaurant_from_token()

    if err:
        return jsonify({"error": err[0]}), err[1]

    status = request.args.get("status")
    supplier = request.args.get("supplier")
    search = request.args.get("search")

    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    conn = get_db_connection()

    cur = conn.cursor(cursor_factory=RealDictCursor)

    try:

        # =====================================================
        # QUERY
        # =====================================================
        query = """
            SELECT
                gh.grn_id,
                gh.order_id,
                gh.status,
                gh.created_at,

                sr.company_name_english,
                sr.company_name_arabic,

                COALESCE(
                    SUM(gi.received_quantity),
                    0
                ) AS received_qty,

                COUNT(gi.grn_item_id) AS total_items

            FROM grn_header gh

            JOIN supplier_registration sr
                ON sr.supplier_id = gh.supplier_id

            LEFT JOIN grn_items gi
                ON gi.grn_id = gh.grn_id

            WHERE gh.restaurant_id = %s
        """

        params = [restaurant_id]

        # =====================================================
        # STATUS FILTER
        # =====================================================
        if status and status != "ALL":

            query += """
                AND gh.status = %s
            """

            params.append(status)

        # =====================================================
        # SUPPLIER FILTER
        # =====================================================
        if supplier and supplier != "ALL":

            query += """
                AND (
                    sr.company_name_english = %s
                    OR sr.company_name_arabic = %s
                )
            """

            params.extend([
                supplier,
                supplier
            ])

        # =====================================================
        # SEARCH FILTER
        # =====================================================
        if search:

            query += """
                AND (
                    CAST(gh.grn_id AS TEXT)
                    ILIKE %s

                    OR CAST(gh.order_id AS TEXT)
                    ILIKE %s

                    OR sr.company_name_english
                    ILIKE %s

                    OR sr.company_name_arabic
                    ILIKE %s
                )
            """

            params.extend([
                f"%{search}%",
                f"%{search}%",
                f"%{search}%",
                f"%{search}%"
            ])

        # =====================================================
        # DATE FILTER
        # =====================================================
        if start_date and end_date:

            query += """
                AND DATE(gh.created_at)
                BETWEEN %s AND %s
            """

            params.extend([
                start_date,
                end_date
            ])

        # =====================================================
        # GROUP + SORT
        # =====================================================
        query += """

            GROUP BY
                gh.grn_id,
                gh.order_id,
                gh.status,
                gh.created_at,
                sr.company_name_english,
                sr.company_name_arabic

            ORDER BY gh.created_at DESC
        """

        # =====================================================
        # EXECUTE
        # =====================================================
        cur.execute(query, tuple(params))

        rows = cur.fetchall()

        # =====================================================
        # PDF BUFFER
        # =====================================================
        buffer = io.BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=20,
            leftMargin=20,
            topMargin=20,
            bottomMargin=20
        )

        styles = getSampleStyleSheet()

        elements = []

        # =====================================================
        # TITLE
        # =====================================================
        elements.append(
            Paragraph(
                "<b>MAHAL GRN REPORT</b>",
                styles["Title"]
            )
        )

        elements.append(Spacer(1, 10))

        # =====================================================
        # FILTER INFO
        # =====================================================
        filter_text = f'''
            <b>Status:</b> {status or "ALL"}
            &nbsp;&nbsp;&nbsp;
            <b>Supplier:</b> {supplier or "ALL"}
            &nbsp;&nbsp;&nbsp;
            <b>Date:</b> {start_date or "-"} to {end_date or "-"}
        '''

        elements.append(
            Paragraph(
                filter_text,
                styles["BodyText"]
            )
        )

        elements.append(Spacer(1, 12))

        # =====================================================
        # TABLE DATA
        # =====================================================
        data = [[
            "GRN",
            "ORDER",
            "SUPPLIER",
            "STATUS",
            "QTY",
            "ITEMS",
            "DATE"
        ]]

        total_qty = 0

        for r in rows:

            total_qty += float(
                r["received_qty"] or 0
            )

            data.append([
                f'GRN-{r["grn_id"]}',
                str(r["order_id"]),
                r["company_name_english"],
                r["status"],
                str(r["received_qty"]),
                str(r["total_items"]),
                str(r["created_at"])[:10]
            ])

        # =====================================================
        # TOTAL ROW
        # =====================================================
        data.append([
            "",
            "",
            "",
            "TOTAL",
            str(total_qty),
            "",
            ""
        ])

        # =====================================================
        # CREATE TABLE
        # =====================================================
        table = Table(
            data,
            colWidths=[
                0.8 * inch,
                1.4 * inch,
                2.2 * inch,
                1.2 * inch,
                0.8 * inch,
                0.8 * inch,
                1.0 * inch
            ]
        )

        # =====================================================
        # TABLE STYLE
        # =====================================================
        table.setStyle(TableStyle([

            # HEADER
            ("BACKGROUND", (0, 0), (-1, 0), colors.orange),

            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),

            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),

            ("FONTSIZE", (0, 0), (-1, -1), 8),

            ("BOTTOMPADDING", (0, 0), (-1, 0), 10),

            # GRID
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),

            # ALIGN
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),

            # BODY BG
            ("BACKGROUND", (0, 1), (-1, -2), colors.whitesmoke),

            # TOTAL ROW
            ("BACKGROUND", (0, -1), (-1, -1), colors.lightgrey),

            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),

        ]))

        elements.append(table)

        elements.append(Spacer(1, 20))

        # =====================================================
        # FOOTER
        # =====================================================
        elements.append(
            Paragraph(
                f"""
                Total GRNs: <b>{len(rows)}</b>
                <br/>
                Total Received Qty:
                <b>{total_qty}</b>
                """,
                styles["BodyText"]
            )
        )

        # =====================================================
        # BUILD PDF
        # =====================================================
        doc.build(elements)

        buffer.seek(0)

        return send_file(
            buffer,
            download_name="grn_report.pdf",
            as_attachment=True
        )

    finally:

        cur.close()
        conn.close()

# =========================================================
# ENTERPRISE INVOICE DETAILS
# =========================================================
@restaurant_reports_bp.route(
    "/restaurant/reports/invoices/<invoice_id>",
    methods=["GET"]
)
def invoice_details(invoice_id):

    restaurant_id, err = get_restaurant_from_token()

    if err:
        return jsonify({
            "error": err[0]
        }), err[1]

    conn = get_db_connection()

    cur = conn.cursor(
        cursor_factory=RealDictCursor
    )

    try:

        # =================================================
        # HEADER
        # =================================================
        cur.execute("""

            SELECT
                ih.*,

                sr.company_name_english
                    AS supplier_name,

                sr.company_name_arabic
                    AS supplier_name_arabic

            FROM invoice_header ih

            JOIN supplier_registration sr
                ON sr.supplier_id =
                    ih.supplier_id

            WHERE ih.invoice_id = %s
            AND ih.restaurant_id = %s

        """, (
            invoice_id,
            restaurant_id
        ))

        invoice = cur.fetchone()

        if not invoice:

            return jsonify({
                "error": "Invoice not found"
            }), 404

        # =================================================
        # ITEMS
        # =================================================
        cur.execute("""

            SELECT
                ii.invoice_item_id,

                ii.product_id,

                ii.product_name_english,

                ii.quantity,

                ii.price_per_unit,

                ii.discount,

                ii.total_amount

            FROM invoice_items ii

            WHERE ii.invoice_id = %s

            ORDER BY ii.invoice_item_id

        """, (invoice_id,))

        items = cur.fetchall()

        return jsonify({
            "invoice": invoice,
            "items": items
        }), 200

    finally:

        cur.close()
        conn.close()


# =========================================================
# INVOICE EXCEL EXPORT
# =========================================================
@restaurant_reports_bp.route(
    "/restaurant/reports/invoices/export",
    methods=["GET"]
)
def export_invoice_excel():

    restaurant_id, err = get_restaurant_from_token()

    if err:
        return jsonify({
            "error": err[0]
        }), err[1]

    status = request.args.get("status")
    supplier = request.args.get("supplier")
    search = request.args.get("search")

    start_date = request.args.get(
        "start_date"
    )

    end_date = request.args.get(
        "end_date"
    )

    conn = get_db_connection()

    cur = conn.cursor(
        cursor_factory=RealDictCursor
    )

    try:

        query = """

            SELECT
                ih.invoice_id,
                ih.invoice_number,
                ih.order_id,

                ih.invoice_date,

                ih.payment_status,

                ih.subtotal_amount,
                ih.tax_amount,
                ih.discount_amount,
                ih.grand_total,

                sr.company_name_english
                    AS supplier_name,

                COUNT(
                    ii.invoice_item_id
                ) AS total_items

            FROM invoice_header ih

            JOIN supplier_registration sr
                ON sr.supplier_id =
                    ih.supplier_id

            LEFT JOIN invoice_items ii
                ON ii.invoice_id =
                    ih.invoice_id

            WHERE ih.restaurant_id = %s

        """

        params = [restaurant_id]

        # =============================================
        # STATUS
        # =============================================
        if status and status != "ALL":

            query += """
                AND ih.payment_status = %s
            """

            params.append(status)

        # =============================================
        # SUPPLIER
        # =============================================
        if supplier and supplier != "ALL":

            query += """
                AND (
                    sr.company_name_english = %s
                    OR sr.company_name_arabic = %s
                )
            """

            params.extend([
                supplier,
                supplier
            ])

        # =============================================
        # SEARCH
        # =============================================
        if search:

            query += """

                AND (

                    CAST(
                        ih.invoice_id
                        AS TEXT
                    ) ILIKE %s

                    OR CAST(
                        ih.order_id
                        AS TEXT
                    ) ILIKE %s

                    OR sr.company_name_english
                        ILIKE %s

                )

            """

            params.extend([
                f"%{search}%",
                f"%{search}%",
                f"%{search}%"
            ])

        # =============================================
        # DATE
        # =============================================
        if start_date and end_date:

            query += """

                AND DATE(
                    ih.invoice_date
                )

                BETWEEN %s AND %s

            """

            params.extend([
                start_date,
                end_date
            ])

        query += """

            GROUP BY
                ih.invoice_id,
                ih.invoice_number,
                ih.order_id,
                ih.invoice_date,
                ih.payment_status,
                ih.subtotal_amount,
                ih.tax_amount,
                ih.discount_amount,
                ih.grand_total,
                sr.company_name_english

            ORDER BY
                ih.invoice_date DESC

        """

        cur.execute(
            query,
            tuple(params)
        )

        rows = cur.fetchall()

        # =============================================
        # EXCEL
        # =============================================
        wb = Workbook()

        ws = wb.active

        ws.title = "Invoice Report"

        headers = [

            "Invoice No",

            "Order ID",

            "Supplier",

            "Status",

            "Subtotal",

            "Tax",

            "Discount",

            "Grand Total",

            "Items",

            "Date"

        ]

        ws.append(headers)

        # =============================================
        # STYLE
        # =============================================
        for cell in ws[1]:

            cell.font = Font(
                bold=True,
                size=12
            )

            cell.alignment = Alignment(
                horizontal="center"
            )

        # =============================================
        # ROWS
        # =============================================
        for r in rows:

            ws.append([

                r["invoice_number"],

                r["order_id"],

                r["supplier_name"],

                r["payment_status"],

                float(
                    r["subtotal_amount"] or 0
                ),

                float(
                    r["tax_amount"] or 0
                ),

                float(
                    r["discount_amount"] or 0
                ),

                float(
                    r["grand_total"] or 0
                ),

                r["total_items"],

                str(
                    r["invoice_date"]
                )[:10]

            ])

        # =============================================
        # WIDTHS
        # =============================================
        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 35
        ws.column_dimensions["D"].width = 18
        ws.column_dimensions["E"].width = 18
        ws.column_dimensions["F"].width = 18
        ws.column_dimensions["G"].width = 18
        ws.column_dimensions["H"].width = 20
        ws.column_dimensions["I"].width = 12
        ws.column_dimensions["J"].width = 18

        output = io.BytesIO()

        wb.save(output)

        output.seek(0)

        return send_file(
            output,
            download_name=
                "invoice_report.xlsx",
            as_attachment=True
        )

    finally:

        cur.close()
        conn.close()

# =========================================================
# INVOICE PDF EXPORT
# =========================================================
@restaurant_reports_bp.route(
    "/restaurant/reports/invoices/pdf",
    methods=["GET"]
)
def export_invoice_pdf():

    restaurant_id, err = get_restaurant_from_token()

    if err:
        return jsonify({
            "error": err[0]
        }), err[1]

    conn = get_db_connection()

    cur = conn.cursor(
        cursor_factory=RealDictCursor
    )

    try:

        cur.execute("""

            SELECT
                ih.invoice_number,
                ih.order_id,

                ih.invoice_date,

                ih.payment_status,

                ih.subtotal_amount,
                ih.tax_amount,
                ih.discount_amount,
                ih.grand_total,

                sr.company_name_english
                    AS supplier_name,

                COUNT(
                    ii.invoice_item_id
                ) AS total_items

            FROM invoice_header ih

            JOIN supplier_registration sr
                ON sr.supplier_id =
                    ih.supplier_id

            LEFT JOIN invoice_items ii
                ON ii.invoice_id =
                    ih.invoice_id

            WHERE ih.restaurant_id = %s

            GROUP BY
                ih.invoice_number,
                ih.order_id,
                ih.invoice_date,
                ih.payment_status,
                ih.subtotal_amount,
                ih.tax_amount,
                ih.discount_amount,
                ih.grand_total,
                sr.company_name_english

            ORDER BY
                ih.invoice_date DESC

        """, (restaurant_id,))

        rows = cur.fetchall()

        buffer = io.BytesIO()

        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=20,
            leftMargin=20,
            topMargin=20,
            bottomMargin=20
        )

        styles = getSampleStyleSheet()

        elements = []

        # =============================================
        # TITLE
        # =============================================
        elements.append(
            Paragraph(
                "<b>MAHAL INVOICE REPORT</b>",
                styles["Title"]
            )
        )

        elements.append(
            Spacer(1, 12)
        )

        # =============================================
        # TABLE
        # =============================================
        data = [[

            "Invoice",

            "Order",

            "Supplier",

            "Status",

            "Grand Total",

            "Items",

            "Date"

        ]]

        total_amount = 0

        for r in rows:

            total_amount += float(
                r["grand_total"] or 0
            )

            data.append([

                r["invoice_number"],

                r["order_id"],

                r["supplier_name"],

                r["payment_status"],

                f'QAR {r["grand_total"]}',

                str(
                    r["total_items"]
                ),

                str(
                    r["invoice_date"]
                )[:10]

            ])

        # =============================================
        # TOTAL ROW
        # =============================================
        data.append([

            "",
            "",
            "",
            "TOTAL",

            f"QAR {total_amount:.2f}",

            "",
            ""

        ])

        table = Table(
            data,
            colWidths=[
                0.8 * inch,
                1.2 * inch,
                2.2 * inch,
                1.0 * inch,
                1.3 * inch,
                0.7 * inch,
                0.8 * inch
            ]
        )

        table.setStyle(TableStyle([

            ("BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.orange),

            ("TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white),

            ("GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey),

            ("FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"),

            ("FONTSIZE",
                (0, 0),
                (-1, -1),
                8),

            ("ALIGN",
                (0, 0),
                (-1, -1),
                "CENTER"),

            ("BACKGROUND",
                (0, 1),
                (-1, -2),
                colors.whitesmoke),

            ("BACKGROUND",
                (0, -1),
                (-1, -1),
                colors.lightgrey),

            ("FONTNAME",
                (0, -1),
                (-1, -1),
                "Helvetica-Bold"),

        ]))

        elements.append(table)

        elements.append(
            Spacer(1, 20)
        )

        elements.append(
            Paragraph(
                f"""

                Total Invoices:
                <b>{len(rows)}</b>

                <br/>

                Total Amount:
                <b>
                    QAR {total_amount:.2f}
                </b>

                """,
                styles["BodyText"]
            )
        )

        doc.build(elements)

        buffer.seek(0)

        return send_file(
            buffer,
            download_name=
                "invoice_report.pdf",
            as_attachment=True
        )

    finally:

        cur.close()
        conn.close()






# =========================================================
# SUPPLIER REPORT EXCEL EXPORT
# =========================================================

@restaurant_reports_bp.route(
    "/restaurant/reports/suppliers/export",
    methods=["GET"]
)
def export_supplier_excel():

    restaurant_id, err = get_restaurant_from_token()

    if err:
        return jsonify({
            "error": err[0]
        }), err[1]

    search = request.args.get("search")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    conn = get_db_connection()

    cur = conn.cursor(
        cursor_factory=RealDictCursor
    )

    try:

        # =====================================================
        # QUERY
        # =====================================================

        query = """

            SELECT
                sr.supplier_id,

                sr.company_name_english,

                sr.company_name_arabic,

                COUNT(
                    DISTINCT oh.order_id
                ) AS total_orders,

                COALESCE(
                    SUM(oh.total_amount),
                    0
                ) AS total_purchase,

                COUNT(
                    CASE
                        WHEN oh.status = 'DELIVERED'
                        THEN 1
                    END
                ) AS delivered_orders,

                COUNT(
                    CASE
                        WHEN oh.status != 'DELIVERED'
                        THEN 1
                    END
                ) AS pending_orders

            FROM supplier_registration sr

            JOIN order_header oh
                ON oh.supplier_id =
                    sr.supplier_id

            WHERE oh.restaurant_id = %s

        """

        params = [restaurant_id]

        # =====================================================
        # SEARCH
        # =====================================================

        if search:

            query += """

                AND (

                    sr.company_name_english
                    ILIKE %s

                    OR sr.company_name_arabic
                    ILIKE %s

                )

            """

            params.extend([
                f"%{search}%",
                f"%{search}%"
            ])

        # =====================================================
        # DATE FILTER
        # =====================================================

        if start_date and end_date:

            query += """

                AND DATE(oh.order_date)
                BETWEEN %s AND %s

            """

            params.extend([
                start_date,
                end_date
            ])

        # =====================================================
        # GROUP
        # =====================================================

        query += """

            GROUP BY

                sr.supplier_id,

                sr.company_name_english,

                sr.company_name_arabic

            ORDER BY
                total_purchase DESC

        """

        cur.execute(
            query,
            tuple(params)
        )

        rows = cur.fetchall()

        # =====================================================
        # EXCEL
        # =====================================================

        wb = Workbook()

        ws = wb.active

        ws.title = "Supplier Report"

        headers = [

            "Supplier",

            "Orders",

            "Delivered",

            "Pending",

            "Total Purchase",

            "Success %"

        ]

        ws.append(headers)

        # =====================================================
        # HEADER STYLE
        # =====================================================

        for cell in ws[1]:

            cell.font = Font(
                bold=True,
                size=12
            )

            cell.alignment = Alignment(
                horizontal="center"
            )

        # =====================================================
        # ROWS
        # =====================================================

        total_purchase = 0

        for r in rows:

            total_orders = int(
                r["total_orders"] or 0
            )

            delivered = int(
                r["delivered_orders"] or 0
            )

            pending = int(
                r["pending_orders"] or 0
            )

            purchase = float(
                r["total_purchase"] or 0
            )

            success = (
                0
                if total_orders == 0
                else round(
                    (delivered / total_orders)
                    * 100
                )
            )

            total_purchase += purchase

            ws.append([

                r["company_name_english"],

                total_orders,

                delivered,

                pending,

                purchase,

                f"{success}%"

            ])

        # =====================================================
        # TOTAL ROW
        # =====================================================

        ws.append([

            "TOTAL",

            "",

            "",

            "",

            total_purchase,

            ""

        ])

        # =====================================================
        # WIDTHS
        # =====================================================

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 22
        ws.column_dimensions["F"].width = 15

        # =====================================================
        # SAVE
        # =====================================================

        output = io.BytesIO()

        wb.save(output)

        output.seek(0)

        return send_file(

            output,

            download_name=
                "supplier_report.xlsx",

            as_attachment=True

        )

    finally:

        cur.close()

        conn.close()


# =========================================================
# SUPPLIER REPORT PDF EXPORT
# =========================================================

@restaurant_reports_bp.route(
    "/restaurant/reports/suppliers/pdf",
    methods=["GET"]
)
def export_supplier_pdf():

    restaurant_id, err = get_restaurant_from_token()

    if err:
        return jsonify({
            "error": err[0]
        }), err[1]

    search = request.args.get("search")
    start_date = request.args.get("start_date")
    end_date = request.args.get("end_date")

    conn = get_db_connection()

    cur = conn.cursor(
        cursor_factory=RealDictCursor
    )

    try:

        # =====================================================
        # QUERY
        # =====================================================

        query = """

            SELECT
                sr.supplier_id,

                sr.company_name_english,

                sr.company_name_arabic,

                COUNT(
                    DISTINCT oh.order_id
                ) AS total_orders,

                COALESCE(
                    SUM(oh.total_amount),
                    0
                ) AS total_purchase,

                COUNT(
                    CASE
                        WHEN oh.status = 'DELIVERED'
                        THEN 1
                    END
                ) AS delivered_orders,

                COUNT(
                    CASE
                        WHEN oh.status != 'DELIVERED'
                        THEN 1
                    END
                ) AS pending_orders

            FROM supplier_registration sr

            JOIN order_header oh
                ON oh.supplier_id =
                    sr.supplier_id

            WHERE oh.restaurant_id = %s

        """

        params = [restaurant_id]

        # =====================================================
        # SEARCH
        # =====================================================

        if search:

            query += """

                AND (

                    sr.company_name_english
                    ILIKE %s

                    OR sr.company_name_arabic
                    ILIKE %s

                )

            """

            params.extend([
                f"%{search}%",
                f"%{search}%"
            ])

        # =====================================================
        # DATE
        # =====================================================

        if start_date and end_date:

            query += """

                AND DATE(oh.order_date)
                BETWEEN %s AND %s

            """

            params.extend([
                start_date,
                end_date
            ])

        # =====================================================
        # GROUP
        # =====================================================

        query += """

            GROUP BY

                sr.supplier_id,

                sr.company_name_english,

                sr.company_name_arabic

            ORDER BY
                total_purchase DESC

        """

        cur.execute(
            query,
            tuple(params)
        )

        rows = cur.fetchall()

        # =====================================================
        # PDF
        # =====================================================

        buffer = io.BytesIO()

        doc = SimpleDocTemplate(

            buffer,

            pagesize=A4,

            rightMargin=20,

            leftMargin=20,

            topMargin=20,

            bottomMargin=20

        )

        styles = getSampleStyleSheet()

        elements = []

        # =====================================================
        # TITLE
        # =====================================================

        elements.append(

            Paragraph(
                "<b>MAHAL SUPPLIER REPORT</b>",
                styles["Title"]
            )

        )

        elements.append(
            Spacer(1, 10)
        )

        # =====================================================
        # FILTER INFO
        # =====================================================

        filter_text = f"""

            <b>Search:</b>
            {search or "ALL"}

            <br/>

            <b>Date:</b>
            {start_date or "-"}
            to
            {end_date or "-"}

        """

        elements.append(

            Paragraph(
                filter_text,
                styles["BodyText"]
            )

        )

        elements.append(
            Spacer(1, 12)
        )

        # =====================================================
        # TABLE
        # =====================================================

        data = [[

            "Supplier",

            "Orders",

            "Delivered",

            "Pending",

            "Purchase",

            "Success"

        ]]

        total_purchase = 0

        total_orders = 0

        for r in rows:

            orders = int(
                r["total_orders"] or 0
            )

            delivered = int(
                r["delivered_orders"] or 0
            )

            pending = int(
                r["pending_orders"] or 0
            )

            purchase = float(
                r["total_purchase"] or 0
            )

            success = (
                0
                if orders == 0
                else round(
                    (delivered / orders)
                    * 100
                )
            )

            total_purchase += purchase

            total_orders += orders

            data.append([

                r["company_name_english"],

                str(orders),

                str(delivered),

                str(pending),

                f"QAR {purchase:.2f}",

                f"{success}%"

            ])

        # =====================================================
        # TOTAL ROW
        # =====================================================

        data.append([

            "TOTAL",

            str(total_orders),

            "",

            "",

            f"QAR {total_purchase:.2f}",

            ""

        ])

        # =====================================================
        # TABLE
        # =====================================================

        table = Table(

            data,

            colWidths=[

                2.5 * inch,

                0.8 * inch,

                1 * inch,

                0.8 * inch,

                1.4 * inch,

                0.8 * inch

            ]

        )

        # =====================================================
        # STYLE
        # =====================================================

        table.setStyle(TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.orange
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                8
            ),

            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "CENTER"
            ),

            (
                "BACKGROUND",
                (0, 1),
                (-1, -2),
                colors.whitesmoke
            ),

            (
                "BACKGROUND",
                (0, -1),
                (-1, -1),
                colors.lightgrey
            ),

            (
                "FONTNAME",
                (0, -1),
                (-1, -1),
                "Helvetica-Bold"
            ),

        ]))

        elements.append(table)

        elements.append(
            Spacer(1, 20)
        )

        # =====================================================
        # FOOTER
        # =====================================================

        elements.append(

            Paragraph(

                f"""

                Total Suppliers:
                <b>{len(rows)}</b>

                <br/>

                Total Purchase:
                <b>
                    QAR {total_purchase:.2f}
                </b>

                """,

                styles["BodyText"]

            )

        )

        # =====================================================
        # BUILD PDF
        # =====================================================

        doc.build(elements)

        buffer.seek(0)

        return send_file(

            buffer,

            download_name=
                "supplier_report.pdf",

            as_attachment=True

        )

    finally:

        cur.close()

        conn.close()

